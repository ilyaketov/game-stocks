"""Генерация xlsx-отчёта с 5 листами: Источник, Свод, Таблица, Отчёт, Риски."""
from datetime import datetime
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ===== Стили (общие для всех листов) =====
HEADER_FILL = PatternFill('solid', start_color='305496')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
THIN = Side(border_style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DATA_FONT = Font(name='Arial', size=10)
DATA_FONT_BOLD = Font(name='Arial', size=10, bold=True)
LEFT = Alignment(horizontal='left', wrap_text=True)
RIGHT = Alignment(horizontal='right')
CENTER = Alignment(horizontal='center')
GROUP_FILL = PatternFill('solid', start_color='D9E1F2')
GROUP_FONT = Font(name='Arial', bold=True, size=10, color='1F4E78')
TOTAL_FILL = PatternFill('solid', start_color='FCE4D6')
TOTAL_FONT = Font(name='Arial', bold=True, size=10)
GRAND_FILL = PatternFill('solid', start_color='305496')
GRAND_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='305496')
SUBTITLE_FONT = Font(name='Arial', italic=True, size=10, color='595959')
SECTION_FONT = Font(name='Arial', bold=True, size=12, color='1F4E78')

YELLOW_FILL = PatternFill('solid', start_color='FFF2CC')
ORANGE_FILL = PatternFill('solid', start_color='F8CBAD')
RED_FILL = PatternFill('solid', start_color='F4B7B7')

HEADERS = [
    "Подразделение", "Игра", "Артикул", "Дата загрузки",
    "Остаток шт. на начало", "Продано, шт.", "Закуплено, шт.", "Остаток шт. на конец",
    "Остаток $ на начало", "Продано, $", "Закуплено, $", "Остаток $ на конец",
    "Дней с последнего закупа", "Движение",
]
WIDTHS = [22, 50, 12, 13, 13, 12, 13, 13, 14, 13, 14, 14, 14, 11]
QTY_COLS = (5, 6, 7, 8)
MONEY_COLS = (9, 10, 11, 12)
SRC_WIDTHS = [42, 42, 14, 22, 18, 30, 10, 10, 14, 12, 14, 22]

MONTHS_RU = {
    1: 'январь', 2: 'февраль', 3: 'март', 4: 'апрель',
    5: 'май', 6: 'июнь', 7: 'июль', 8: 'август',
    9: 'сентябрь', 10: 'октябрь', 11: 'ноябрь', 12: 'декабрь',
}
MONTHS_RU_GENITIVE = {  # «без движения в марте»
    1: 'январе', 2: 'феврале', 3: 'марте', 4: 'апреле',
    5: 'мае', 6: 'июне', 7: 'июле', 8: 'августе',
    9: 'сентябре', 10: 'октябре', 11: 'ноябре', 12: 'декабре',
}


def _get_age_fill(days):
    if days is None:
        return None
    if days >= 90:
        return RED_FILL
    if days >= 60:
        return ORANGE_FILL
    if days >= 30:
        return YELLOW_FILL
    return None


def _apply_data_format(ws, row_idx, col_idx, font=None):
    c = ws.cell(row=row_idx, column=col_idx)
    c.font = font or DATA_FONT
    c.border = BORDER
    if col_idx in (1, 2, 3):
        c.alignment = LEFT
    elif col_idx == 4:
        c.number_format = 'DD.MM.YYYY'
        c.alignment = CENTER
    elif col_idx in QTY_COLS:
        c.number_format = '#,##0;(#,##0);-'
        c.alignment = RIGHT
    elif col_idx in MONEY_COLS:
        c.number_format = '$#,##0.00;($#,##0.00);-'
        c.alignment = RIGHT
    elif col_idx == 13:
        c.number_format = '0;;-'
        c.alignment = CENTER
    elif col_idx == 14:
        c.alignment = CENTER


def _cell_visible_length(cell):
    v = cell.value
    if v is None:
        return 0
    if isinstance(v, str):
        if v.startswith('='):
            return 0
        return len(v)
    if isinstance(v, (int, float)):
        nf = cell.number_format or ''
        if '$' in nf and '0.00' in nf:
            return len(f"${v:,.2f}")
        if '$' in nf:
            return len(f"${v:,.0f}")
        if '0.0%' in nf:
            return len(f"{v * 100:.1f}%")
        if '#,##0' in nf:
            return len(f"{v:,.0f}")
        return len(str(v))
    return len(str(v))


def write_report(period_start, period_end, flat, source_bytes, output_path=None):
    """
    Генерирует xlsx-отчёт.

    Параметры:
      period_start, period_end — границы периода (datetime).
      flat — список словарей-строк плоской таблицы (см. calculator.build_flat).
      source_bytes — bytes/BytesIO с исходным xlsx для листа «Источник» (или None).
      output_path — путь для сохранения. Если None, отчёт возвращается как BytesIO.

    Возвращает: путь к файлу (если output_path задан) или BytesIO.
    """
    from io import BytesIO
    month_word = MONTHS_RU[period_end.month]
    month_word_genitive = MONTHS_RU_GENITIVE[period_end.month]
    year = period_end.year

    wb = Workbook()

    # ===== Лист 1: Источник =====
    ws_src = wb.active
    ws_src.title = "Источник"
    if source_bytes is not None:
        from io import BytesIO as _BytesIO
        if isinstance(source_bytes, (bytes, bytearray)):
            src_stream = _BytesIO(source_bytes)
        else:
            src_stream = source_bytes
            try:
                src_stream.seek(0)
            except Exception:
                pass
        try:
            src_wb = load_workbook(src_stream, data_only=True)
            src_ws = src_wb[src_wb.sheetnames[0]]
            for row in src_ws.iter_rows(values_only=True):
                ws_src.append(row)
            if ws_src['A1'].value:
                ws_src['A1'].font = TITLE_FONT
            if ws_src['A2'].value:
                ws_src['A2'].font = Font(name='Arial', bold=True, size=11, color='595959')
            for col_idx in range(1, 13):
                c = ws_src.cell(row=4, column=col_idx)
                if c.value:
                    c.fill = HEADER_FILL
                    c.font = HEADER_FONT
                    c.alignment = HEADER_ALIGN
                    c.border = BORDER
            for i, w in enumerate(SRC_WIDTHS, 1):
                ws_src.column_dimensions[get_column_letter(i)].width = w
            ws_src.row_dimensions[4].height = 30
            ws_src.freeze_panes = 'A5'
        except Exception as e:
            ws_src['A1'] = f"Не удалось прочитать исходник: {e}"
            ws_src['A1'].font = Font(name='Arial', italic=True, size=10, color='808080')
    else:
        ws_src['A1'] = "Источник недоступен"
        ws_src['A1'].font = Font(name='Arial', italic=True, size=10, color='808080')

    # ===== Лист 2: Свод =====
    ws_svod = wb.create_sheet("Свод")

    # ===== Лист 3: Таблица =====
    ws1 = wb.create_sheet("Таблица")
    ws1.append([f"Таблица по остаткам игровых ключей за {month_word} {year}"])
    ws1['A1'].font = TITLE_FONT
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws1.append([f"Период: {period_start.strftime('%d.%m.%Y')} — {period_end.strftime('%d.%m.%Y')}"])
    ws1['A2'].font = SUBTITLE_FONT
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
    ws1.append([])

    T1_HEADER_ROW = 4
    ws1.append(HEADERS)
    flat_sorted_t1 = sorted(flat, key=lambda r: (r['store'], r['game'] or '', r['sku']))
    for r in flat_sorted_t1:
        movement_text = "есть" if r['has_movement_in_period'] else "нет"
        ws1.append([
            r['store'], r['game'], r['sku'], r['upload_date'],
            r['start_qty'], r['sold_qty'], r['bought_qty'], r['end_qty'],
            r['start_val'], r['sold_val'], r['bought_val'], r['end_val'],
            r['days_since_bill'], movement_text,
        ])

    for col_idx in range(1, len(HEADERS) + 1):
        c = ws1.cell(row=T1_HEADER_ROW, column=col_idx)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        c.border = BORDER

    T1_DATA_FIRST = T1_HEADER_ROW + 1
    T1_DATA_LAST = T1_HEADER_ROW + len(flat_sorted_t1)

    for idx, r in enumerate(flat_sorted_t1):
        row_idx = T1_DATA_FIRST + idx
        is_bold = (r['end_qty'] > 0) and (not r['has_movement_in_period'])
        fnt = DATA_FONT_BOLD if is_bold else DATA_FONT
        fill = _get_age_fill(r['days_since_bill']) if r['end_qty'] > 0 else None
        for col_idx in range(1, len(HEADERS) + 1):
            _apply_data_format(ws1, row_idx, col_idx, font=fnt)
            if fill:
                ws1.cell(row=row_idx, column=col_idx).fill = fill

    for i, w in enumerate(WIDTHS, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.row_dimensions[T1_HEADER_ROW].height = 38
    ws1.freeze_panes = f'D{T1_DATA_FIRST}'
    ws1.auto_filter.ref = f"A{T1_HEADER_ROW}:{get_column_letter(len(HEADERS))}{T1_DATA_LAST}"

    sku_store_to_row = {(r['sku'], r['store']): T1_DATA_FIRST + idx for idx, r in enumerate(flat_sorted_t1)}

    # ===== Лист 4: Отчёт =====
    ws2 = wb.create_sheet("Отчёт")
    flat_sorted_t2 = sorted(flat, key=lambda r: (r['store'], -r['end_val'], r['game'] or '', r['sku']))

    ws2.append([f"Отчёт по остаткам игровых ключей за {month_word} {year}"])
    ws2['A1'].font = TITLE_FONT
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws2.append([f"Период: {period_start.strftime('%d.%m.%Y')} — {period_end.strftime('%d.%m.%Y')}"])
    ws2['A2'].font = SUBTITLE_FONT
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
    ws2.append([])

    header_row = 4
    ws2.append(HEADERS)
    for col_idx in range(1, len(HEADERS) + 1):
        c = ws2.cell(row=header_row, column=col_idx)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        c.border = BORDER
    ws2.row_dimensions[header_row].height = 38

    current_store = None
    current_group_start = None
    row_cursor = header_row + 1
    group_total_rows = []

    def close_group(rc, store):
        ws2.cell(row=rc, column=1, value=f"Итого по {store}").font = TOTAL_FONT
        for col_idx in list(QTY_COLS) + list(MONEY_COLS):
            col_letter = get_column_letter(col_idx)
            ws2.cell(row=rc, column=col_idx,
                     value=f"=SUM({col_letter}{current_group_start}:{col_letter}{rc - 1})")
        for col_idx in range(1, len(HEADERS) + 1):
            c = ws2.cell(row=rc, column=col_idx)
            c.fill = TOTAL_FILL
            if col_idx > 1:
                c.font = TOTAL_FONT
            c.border = BORDER
            if col_idx in QTY_COLS:
                c.number_format = '#,##0;(#,##0);-'
                c.alignment = RIGHT
            elif col_idx in MONEY_COLS:
                c.number_format = '$#,##0.00;($#,##0.00);-'
                c.alignment = RIGHT
        group_total_rows.append(rc)

    for r in flat_sorted_t2:
        if r['store'] != current_store:
            if current_store is not None:
                close_group(row_cursor, current_store)
                row_cursor += 1
                ws2.append([])
                row_cursor += 1
            current_store = r['store']
            ws2.cell(row=row_cursor, column=1, value=f"▼ {current_store}").font = GROUP_FONT
            for col_idx in range(1, len(HEADERS) + 1):
                c = ws2.cell(row=row_cursor, column=col_idx)
                c.fill = GROUP_FILL
                if col_idx > 1:
                    c.font = GROUP_FONT
                c.border = BORDER
            ws2.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=len(HEADERS))
            row_cursor += 1
            current_group_start = row_cursor

        movement_text = "есть" if r['has_movement_in_period'] else "нет"
        is_bold = (r['end_qty'] > 0) and (not r['has_movement_in_period'])
        fnt = DATA_FONT_BOLD if is_bold else DATA_FONT
        fill = _get_age_fill(r['days_since_bill']) if r['end_qty'] > 0 else None

        ws2.cell(row=row_cursor, column=1, value=r['store'])
        ws2.cell(row=row_cursor, column=2, value=r['game'])
        ws2.cell(row=row_cursor, column=3, value=r['sku'])
        ws2.cell(row=row_cursor, column=4, value=r['upload_date'])
        ws2.cell(row=row_cursor, column=5, value=r['start_qty'])
        ws2.cell(row=row_cursor, column=6, value=r['sold_qty'])
        ws2.cell(row=row_cursor, column=7, value=r['bought_qty'])
        ws2.cell(row=row_cursor, column=8, value=r['end_qty'])
        ws2.cell(row=row_cursor, column=9, value=r['start_val'])
        ws2.cell(row=row_cursor, column=10, value=r['sold_val'])
        ws2.cell(row=row_cursor, column=11, value=r['bought_val'])
        ws2.cell(row=row_cursor, column=12, value=r['end_val'])
        ws2.cell(row=row_cursor, column=13, value=r['days_since_bill'])
        ws2.cell(row=row_cursor, column=14, value=movement_text)
        for col_idx in range(1, len(HEADERS) + 1):
            _apply_data_format(ws2, row_cursor, col_idx, font=fnt)
            if fill:
                ws2.cell(row=row_cursor, column=col_idx).fill = fill
        row_cursor += 1

    if current_store is not None:
        close_group(row_cursor, current_store)
        row_cursor += 1

    row_cursor += 1
    ws2.cell(row=row_cursor, column=1, value="ИТОГО ПО ВСЕМ ПОДРАЗДЕЛЕНИЯМ")
    for col_idx in list(QTY_COLS) + list(MONEY_COLS):
        col_letter = get_column_letter(col_idx)
        parts = [f"{col_letter}{tr}" for tr in group_total_rows]
        formula = "=" + "+".join(parts) if parts else "=0"
        ws2.cell(row=row_cursor, column=col_idx, value=formula)
    for col_idx in range(1, len(HEADERS) + 1):
        c = ws2.cell(row=row_cursor, column=col_idx)
        c.fill = GRAND_FILL
        c.font = GRAND_FONT
        c.border = BORDER
        c.alignment = Alignment(horizontal='center' if col_idx == 1 else 'right', vertical='center')
        if col_idx in QTY_COLS:
            c.number_format = '#,##0;(#,##0);-'
        elif col_idx in MONEY_COLS:
            c.number_format = '$#,##0.00;($#,##0.00);-'

    for i, w in enumerate(WIDTHS, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = f'A{header_row + 1}'

    # ===== Заполняем "Свод" =====
    ws_svod['A1'] = f"Сводная по подразделениям. {month_word.capitalize()} {year}"
    ws_svod['A1'].font = TITLE_FONT
    ws_svod.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws_svod['A2'] = ""

    pivot_headers = [
        "Подразделение", "Остаток шт. на начало", "Продано, шт.", "Закуплено, шт.", "Остаток шт. на конец",
        "Остаток $ на начало", "Продано, $", "Закуплено, $", "Остаток $ на конец",
    ]
    phr = 3
    for col_idx, h in enumerate(pivot_headers, 1):
        c = ws_svod.cell(row=phr, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        c.border = BORDER
    ws_svod.row_dimensions[phr].height = 38

    stores = sorted({r['store'] for r in flat})
    T1_FIRST = T1_DATA_FIRST
    T1_LAST = T1_DATA_LAST
    range_a = f"Таблица!$A${T1_FIRST}:$A${T1_LAST}"
    src_cols_map = {2: 'E', 3: 'F', 4: 'G', 5: 'H', 6: 'I', 7: 'J', 8: 'K', 9: 'L'}
    start_data = phr + 1
    for i, store in enumerate(stores):
        r = start_data + i
        ws_svod.cell(row=r, column=1, value=store)
        ws_svod.cell(row=r, column=1).font = DATA_FONT
        ws_svod.cell(row=r, column=1).alignment = LEFT
        ws_svod.cell(row=r, column=1).border = BORDER
        for col_idx, src_col in src_cols_map.items():
            rng = f"Таблица!${src_col}${T1_FIRST}:${src_col}${T1_LAST}"
            ws_svod.cell(row=r, column=col_idx, value=f"=SUMIF({range_a},$A{r},{rng})")
            c = ws_svod.cell(row=r, column=col_idx)
            c.font = DATA_FONT
            c.border = BORDER
            c.alignment = RIGHT
            if col_idx in (2, 3, 4, 5):
                c.number_format = '#,##0;(#,##0);-'
            else:
                c.number_format = '$#,##0.00;($#,##0.00);-'

    total_row = start_data + len(stores)
    ws_svod.cell(row=total_row, column=1, value="ИТОГО")
    for col_idx in range(2, 10):
        col_letter = get_column_letter(col_idx)
        ws_svod.cell(row=total_row, column=col_idx,
                     value=f"=SUM({col_letter}{start_data}:{col_letter}{total_row - 1})")
    for col_idx in range(1, 10):
        c = ws_svod.cell(row=total_row, column=col_idx)
        c.fill = GRAND_FILL
        c.font = GRAND_FONT
        c.border = BORDER
        c.alignment = Alignment(horizontal='center' if col_idx == 1 else 'right', vertical='center')
        if col_idx in (2, 3, 4, 5):
            c.number_format = '#,##0;(#,##0);-'
        elif col_idx >= 6:
            c.number_format = '$#,##0.00;($#,##0.00);-'

    pivot_widths = [24, 15, 13, 14, 15, 17, 14, 15, 17]
    for i, w in enumerate(pivot_widths, 1):
        ws_svod.column_dimensions[get_column_letter(i)].width = w
    ws_svod.freeze_panes = f'B{phr + 1}'

    # ===== Лист 5: Риски =====
    ws3 = wb.create_sheet("Риски")
    T = "Таблица"
    T_DAYS = f"{T}!$M${T1_FIRST}:$M${T1_LAST}"
    T_END_QTY = f"{T}!$H${T1_FIRST}:$H${T1_LAST}"
    T_END_VAL = f"{T}!$L${T1_FIRST}:$L${T1_LAST}"
    T_STORE = f"{T}!$A${T1_FIRST}:$A${T1_LAST}"

    ws3['A1'] = f"Стареющий сток на {period_end.strftime('%d.%m.%Y')}"
    ws3['A1'].font = TITLE_FONT
    ws3.merge_cells('A1:I1')

    ws3['A2'] = (f"Игры в стоке: цвет — по дням с последнего закупа; "
                 f"жирный шрифт — без движения в {month_word_genitive}")
    ws3['A2'].font = SUBTITLE_FONT
    ws3.merge_cells('A2:I2')

    # === Категории риска ===
    ws3['A4'] = "Категории риска"
    ws3['A4'].font = SECTION_FONT
    ws3.merge_cells('A4:K4')

    cat_hdr_row = 5
    cat_hdrs = ["Категория", "Описание", "Позиций", "Шт. остаток", "$ остаток"]
    for col, h in enumerate(cat_hdrs, 1):
        c = ws3.cell(row=cat_hdr_row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        c.border = BORDER
    ws3.row_dimensions[cat_hdr_row].height = 25

    cat_data = [
        ("30–60 дней", "Лежит 1–2 месяца", YELLOW_FILL, ">=30", "<60"),
        ("60–90 дней", "Лежит 2–3 месяца", ORANGE_FILL, ">=60", "<90"),
        ("90+ дней", "С самого старта учёта", RED_FILL, ">=90", None),
    ]
    for i, (label, desc, fill, lo_op, hi_op) in enumerate(cat_data):
        r = cat_hdr_row + 1 + i
        ws3.cell(row=r, column=1, value=label).fill = fill
        ws3.cell(row=r, column=2, value=desc)
        if hi_op:
            ws3.cell(row=r, column=3, value=f'=COUNTIFS({T_END_QTY},">0",{T_DAYS},"{lo_op}",{T_DAYS},"{hi_op}")')
            ws3.cell(row=r, column=4, value=f'=SUMIFS({T_END_QTY},{T_END_QTY},">0",{T_DAYS},"{lo_op}",{T_DAYS},"{hi_op}")')
            ws3.cell(row=r, column=5, value=f'=SUMIFS({T_END_VAL},{T_END_QTY},">0",{T_DAYS},"{lo_op}",{T_DAYS},"{hi_op}")')
        else:
            ws3.cell(row=r, column=3, value=f'=COUNTIFS({T_END_QTY},">0",{T_DAYS},"{lo_op}")')
            ws3.cell(row=r, column=4, value=f'=SUMIFS({T_END_QTY},{T_END_QTY},">0",{T_DAYS},"{lo_op}")')
            ws3.cell(row=r, column=5, value=f'=SUMIFS({T_END_VAL},{T_END_QTY},">0",{T_DAYS},"{lo_op}")')
        for col in range(1, 6):
            c = ws3.cell(row=r, column=col)
            c.font = DATA_FONT
            c.border = BORDER
            if col == 1:
                c.font = Font(name='Arial', bold=True, size=10)
            elif col == 2:
                c.alignment = Alignment(horizontal='left')
            elif col == 3:
                c.number_format = '#,##0'
                c.alignment = RIGHT
            elif col == 4:
                c.number_format = '#,##0'
                c.alignment = RIGHT
            elif col == 5:
                c.number_format = '$#,##0.00'
                c.alignment = RIGHT

    itog_row = cat_hdr_row + 1 + len(cat_data)
    ws3.cell(row=itog_row, column=1, value="Всего стареющего")
    ws3.cell(row=itog_row, column=2, value="(остаток > 0, ≥ 30 дней с закупа)")
    ws3.cell(row=itog_row, column=3, value=f"=SUM(C{cat_hdr_row + 1}:C{itog_row - 1})")
    ws3.cell(row=itog_row, column=4, value=f"=SUM(D{cat_hdr_row + 1}:D{itog_row - 1})")
    ws3.cell(row=itog_row, column=5, value=f"=SUM(E{cat_hdr_row + 1}:E{itog_row - 1})")
    for col in range(1, 6):
        c = ws3.cell(row=itog_row, column=col)
        c.fill = GRAND_FILL
        c.font = GRAND_FONT
        c.border = BORDER
        if col == 3:
            c.number_format = '#,##0'
            c.alignment = RIGHT
        elif col == 4:
            c.number_format = '#,##0'
            c.alignment = RIGHT
        elif col == 5:
            c.number_format = '$#,##0.00'
            c.alignment = RIGHT

    # === По подразделениям ===
    sec_row = itog_row + 3
    ws3.cell(row=sec_row, column=1, value="Стареющий сток по подразделениям").font = SECTION_FONT
    ws3.merge_cells(start_row=sec_row, start_column=1, end_row=sec_row, end_column=11)

    def calc_risk_pct(store):
        problem = sum(r['end_val'] for r in flat
                      if r['store'] == store and r['end_qty'] > 0
                      and r['days_since_bill'] is not None and r['days_since_bill'] >= 30)
        total = sum(r['end_val'] for r in flat if r['store'] == store and r['end_qty'] > 0)
        return (problem / total) if total > 0 else 0

    stores_sorted_by_risk = sorted(stores, key=lambda s: -calc_risk_pct(s))

    table_row = sec_row + 1
    hdr2 = ["Подразделение",
            "30-60 дн., шт.", "30-60 дн., $",
            "60-90 дн., шт.", "60-90 дн., $",
            "90+ дн., шт.", "90+ дн., $",
            "Всего стареющ., шт.", "Всего стареющ., $",
            "Остаток магазина, $", "Доля риска"]
    for col, h in enumerate(hdr2, 1):
        c = ws3.cell(row=table_row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        c.border = BORDER
    ws3.row_dimensions[table_row].height = 32

    for i, store in enumerate(stores_sorted_by_risk):
        r = table_row + 1 + i
        ws3.cell(row=r, column=1, value=store)
        ws3.cell(row=r, column=2, value=f'=SUMIFS({T_END_QTY},{T_STORE},$A{r},{T_END_QTY},">0",{T_DAYS},">=30",{T_DAYS},"<60")')
        ws3.cell(row=r, column=3, value=f'=SUMIFS({T_END_VAL},{T_STORE},$A{r},{T_END_QTY},">0",{T_DAYS},">=30",{T_DAYS},"<60")')
        ws3.cell(row=r, column=4, value=f'=SUMIFS({T_END_QTY},{T_STORE},$A{r},{T_END_QTY},">0",{T_DAYS},">=60",{T_DAYS},"<90")')
        ws3.cell(row=r, column=5, value=f'=SUMIFS({T_END_VAL},{T_STORE},$A{r},{T_END_QTY},">0",{T_DAYS},">=60",{T_DAYS},"<90")')
        ws3.cell(row=r, column=6, value=f'=SUMIFS({T_END_QTY},{T_STORE},$A{r},{T_END_QTY},">0",{T_DAYS},">=90")')
        ws3.cell(row=r, column=7, value=f'=SUMIFS({T_END_VAL},{T_STORE},$A{r},{T_END_QTY},">0",{T_DAYS},">=90")')
        ws3.cell(row=r, column=8, value=f"=B{r}+D{r}+F{r}")
        ws3.cell(row=r, column=9, value=f"=C{r}+E{r}+G{r}")
        ws3.cell(row=r, column=10, value=f'=SUMIFS({T_END_VAL},{T_STORE},$A{r},{T_END_QTY},">0")')
        ws3.cell(row=r, column=11, value=f'=IFERROR(I{r}/J{r},0)')
        for col in range(1, 12):
            c = ws3.cell(row=r, column=col)
            c.font = DATA_FONT
            c.border = BORDER
            if col == 1:
                c.alignment = LEFT
            elif col == 11:
                c.number_format = '0.0%'
                c.alignment = RIGHT
            elif col in (3, 5, 7, 9, 10):
                c.number_format = '$#,##0.00'
                c.alignment = RIGHT
            else:
                c.number_format = '#,##0;(#,##0);-'
                c.alignment = RIGHT
        ws3.cell(row=r, column=2).fill = YELLOW_FILL
        ws3.cell(row=r, column=3).fill = YELLOW_FILL
        ws3.cell(row=r, column=4).fill = ORANGE_FILL
        ws3.cell(row=r, column=5).fill = ORANGE_FILL
        ws3.cell(row=r, column=6).fill = RED_FILL
        ws3.cell(row=r, column=7).fill = RED_FILL

    stores_table_end = table_row + len(stores_sorted_by_risk)

    # === Топ-15 ===
    top_row = stores_table_end + 3
    ws3.cell(row=top_row, column=1, value="Топ-15 «мёртвых» позиций (90+ дней, не продавалось)").font = SECTION_FONT
    ws3.merge_cells(start_row=top_row, start_column=1, end_row=top_row, end_column=11)

    problem_rows_data = [r for r in flat
                         if r['days_since_bill'] is not None and r['days_since_bill'] >= 90
                         and r['sold_qty'] == 0 and r['end_qty'] > 0]
    problem_rows_data.sort(key=lambda r: -r['end_val'])
    top15 = problem_rows_data[:15]

    top_hdr_row = top_row + 1
    top_hdr = ["№", "Дней", "Подразделение", "Игра", "Артикул", "Шт. в стоке", "Сумма, $"]
    for i, h in enumerate(top_hdr, 1):
        c = ws3.cell(row=top_hdr_row, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        c.border = BORDER
    ws3.row_dimensions[top_hdr_row].height = 25

    if not top15:
        rr = top_hdr_row + 1
        ws3.cell(row=rr, column=1, value="Нет позиций с возрастом 90+ дней в этом периоде")
        ws3.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=7)
        c = ws3.cell(row=rr, column=1)
        c.font = Font(name='Arial', italic=True, size=10, color='808080')
        c.alignment = CENTER
        c.border = BORDER
        top_end = rr
    else:
        for i, r in enumerate(top15):
            rr = top_hdr_row + 1 + i
            src_r = sku_store_to_row.get((r['sku'], r['store']))
            ws3.cell(row=rr, column=1, value=i + 1)
            ws3.cell(row=rr, column=2, value=f"=Таблица!M{src_r}" if src_r else r['days_since_bill'])
            ws3.cell(row=rr, column=3, value=f"=Таблица!A{src_r}" if src_r else r['store'])
            ws3.cell(row=rr, column=4, value=f"=Таблица!B{src_r}" if src_r else r['game'])
            ws3.cell(row=rr, column=5, value=f"=Таблица!C{src_r}" if src_r else r['sku'])
            ws3.cell(row=rr, column=6, value=f"=Таблица!H{src_r}" if src_r else r['end_qty'])
            ws3.cell(row=rr, column=7, value=f"=Таблица!L{src_r}" if src_r else r['end_val'])
            for col in range(1, 8):
                c = ws3.cell(row=rr, column=col)
                c.font = DATA_FONT
                c.border = BORDER
                if col in (1, 2, 5):
                    c.alignment = CENTER
                elif col in (3, 4):
                    c.alignment = LEFT
                elif col == 6:
                    c.number_format = '#,##0'
                    c.alignment = RIGHT
                elif col == 7:
                    c.number_format = '$#,##0.00'
                    c.alignment = RIGHT
            ws3.cell(row=rr, column=2).fill = RED_FILL
        top_end = top_hdr_row + len(top15)

    # === Выводы ===
    fact_row = top_end + 3
    ws3.cell(row=fact_row, column=1, value="Выводы").font = SECTION_FONT
    ws3.merge_cells(start_row=fact_row, start_column=1, end_row=fact_row, end_column=11)

    never_sold_count = sum(1 for r in flat if not r['has_any_sale_ever'] and r['end_qty'] > 0)
    no_movement_count = sum(1 for r in flat if not r['has_movement_in_period'] and r['end_qty'] > 0)
    has_top15 = len(top15) > 0

    facts = [
        f'="• Стареющий сток (≥30 дней без продаж и остаток > 0): "&TEXT(D{itog_row},"#,##0")&" шт. на сумму "&TEXT(E{itog_row},"$#,##0.00")&" в "&TEXT(C{itog_row},"#,##0")&" позициях."',
        f'="• Из них 90+ дней (с самого старта учёта): "&TEXT(D{cat_hdr_row + 3},"#,##0")&" шт. / "&TEXT(E{cat_hdr_row + 3},"$#,##0.00")&" в "&TEXT(C{cat_hdr_row + 3},"#,##0")&" позициях."',
        f'="• Лидер по объёму стареющего стока: "&A{table_row + 1}&" — "&TEXT(I{table_row + 1},"$#,##0.00")&" ("&TEXT(K{table_row + 1},"0.0%")&" от остатка магазина)."',
        f'="• Без движения в {month_word_genitive} (выделено жирным): {no_movement_count} позиций; никогда не продавалось за весь период учёта: {never_sold_count} позиций."',
    ]
    if has_top15:
        facts.append(
            f'="• Самая «мёртвая» позиция: "&D{top_hdr_row + 1}&" в "&C{top_hdr_row + 1}&" — "&TEXT(F{top_hdr_row + 1},"#,##0")&" шт. на "&TEXT(G{top_hdr_row + 1},"$#,##0.00")&"."'
        )
    else:
        facts.append('="• Позиций с возрастом 90+ дней в этом периоде ещё нет."')

    for i, f_formula in enumerate(facts):
        cr = fact_row + 1 + i
        ws3.cell(row=cr, column=1, value=f_formula)
        ws3.cell(row=cr, column=1).font = Font(name='Arial', size=10)
        ws3.cell(row=cr, column=1).alignment = Alignment(horizontal='left', wrap_text=True, vertical='top')
        ws3.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=11)
        ws3.row_dimensions[cr].height = 20

    # === Автоподбор ширин ===
    merged_ranges = list(ws3.merged_cells.ranges)

    def is_in_merge(cell):
        for rng in merged_ranges:
            if cell.coordinate in rng:
                return True
        return False

    max_col_in_use = 11
    max_lengths = [0] * (max_col_in_use + 1)
    for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, min_col=1, max_col=max_col_in_use):
        for cell in row:
            if is_in_merge(cell):
                continue
            L = _cell_visible_length(cell)
            if L > max_lengths[cell.column]:
                max_lengths[cell.column] = L

    for c in range(1, max_col_in_use + 1):
        width = min(max(max_lengths[c] + 2, 11), 36)
        ws3.column_dimensions[get_column_letter(c)].width = width

    ws3.column_dimensions['D'].width = max(ws3.column_dimensions['D'].width or 0, 38)
    ws3.sheet_view.showGridLines = False

    # Порядок листов
    wb._sheets = [ws_src, ws_svod, ws1, ws2, ws3]

    if output_path is None:
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return str(output_path)
