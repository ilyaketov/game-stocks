"""Парсинг QuickBooks Inventory Report xlsx в плоский список транзакций."""
from datetime import datetime, date
from collections import defaultdict
from pathlib import Path

import openpyxl


def _parse_date(d):
    if isinstance(d, datetime):
        return d
    if isinstance(d, date):
        return datetime.combine(d, datetime.min.time())
    if isinstance(d, str):
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d.%m.%Y'):
            try:
                return datetime.strptime(d, fmt)
            except ValueError:
                continue
    return None


def parse_inventory_report(path_or_file):
    """
    Парсит xlsx-выгрузку в иерархическом формате QuickBooks Inventory Report.

    path_or_file может быть строкой/Path (локальный файл) или file-like объектом
    (например, BytesIO от st.file_uploader).

    Возвращает кортеж (transactions, sku_to_game), где:
      transactions: dict[(sku, store)] -> list[(dt, ttype, qty, cost, on_hand, asset)]
      sku_to_game:  dict[sku] -> название игры

    Игнорирует тип Inventory Starting Value (служебная нулевая транзакция начала учёта).
    """
    wb = openpyxl.load_workbook(path_or_file, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    current_game = None
    current_sku = None
    transactions = defaultdict(list)
    sku_to_game = {}

    for row in rows[4:]:
        if len(row) < 12:
            continue
        a, b, c, d, e, f, g, h, i, j, k, l = row[:12]

        # Заголовок игры или SKU (только колонка A заполнена, нет даты, нет типа)
        if a and not (isinstance(a, str) and a.startswith('Total')) and b is None:
            if str(a).isdigit():
                current_sku = str(a)
                if current_game and current_sku not in sku_to_game:
                    sku_to_game[current_sku] = current_game
            else:
                current_game = a
                current_sku = None
            continue

        # Транзакция
        if b and c and d and current_sku:
            if d == 'Inventory Starting Value':
                continue  # служебная, игнорируем

            store = l or 'UNKNOWN'
            dt = _parse_date(c)
            if dt is None:
                continue

            qty = g or 0
            cost = i or 0
            on_hand = j or 0
            asset = k or 0

            transactions[(current_sku, store)].append((dt, d, qty, cost, on_hand, asset))

    # Стабильная сортировка по дате (порядок строк сохраняется при равных датах)
    for key in transactions:
        transactions[key].sort(key=lambda x: x[0])

    return dict(transactions), sku_to_game


def get_period_range(transactions):
    """Возвращает (min_date, max_date) по всем транзакциям."""
    all_dates = [t[0] for txs in transactions.values() for t in txs]
    if not all_dates:
        return None, None
    return min(all_dates), max(all_dates)


def get_months_in_data(transactions):
    """Возвращает список (year, month) календарных месяцев, представленных в данных."""
    months = set()
    for txs in transactions.values():
        for t in txs:
            months.add((t[0].year, t[0].month))
    return sorted(months)
